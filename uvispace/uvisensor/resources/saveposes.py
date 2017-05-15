#!/usr/bin/env python
"""
Auxiliary program for reading, analyzing and writing data of poses and time.

This module allows:
-Read data of poses with their respective time, from a spreadsheet.
-Read data directly from an array of float32 Nx4.
-Analyze data by calculating differential values of position, time and lenght
 displaced and average speed of UGV
-Save data in spreadsheet and textfile.
-Save final data in master sheet.
"""
# Standard libraries
import math
import numpy as np
import sys
import time
#Excel read/write library
from openpyxl import load_workbook
from openpyxl import Workbook

def read_data(filename_spreadsheet="12_05_2017_1433-L1-R1.xlsx"):
    """
    It allows to read poses and time of spreadsheet to analyze or save them.

    :param filename_spreadsheet: name of spreadsheet that contain the data to
    be read.
    """
    try:
        wb = load_workbook(filename_spreadsheet)
    except:
        wb = Workbook()
    ws = wb.active
    #Initialization of matrices for data.
    data = np.array([0, 0, 0, 0]).astype(np.float32)
    new_data = np.array([0, 0, 0, 0]).astype(np.float32)
    #The first row is the header, and the second is a set of zeros (already
    #initialized in the matrix). Begins to read row 3.
    row = 3
    #Number of columns in the matrix.
    cols = data.shape[0]
    #Loop for reading data.
    current_row_data = True
    while current_row_data:
        element = ws.cell(column=1, row=row).value
        if element == None:
            current_row_data = False
        else:
            for y in range (0, cols):
                element = ws.cell(column=y+1, row=row).value
                new_data[y] = element
            data = np.vstack([data, new_data])
            row +=1
    #Call to save and analyze data.
    save_data(data, analyze=True)

def save_data(data, analyze=False):
    """
    Receives poses and time of matrix to analyze and/or save them.

    :param data: Matrix of floats32 with data.
    :param analyze: Boolean that allows analyze or not the data.
    """
    #Get the SP values from the user.
    time.sleep(0.2)
    #TODO Try, except correct value.
    sp_left = input("Introduce value of sp_left between 0 and 255\n")
    sp_right = input("Introduce value of sp_left between 0 and 255\n")
    #Call for data analysis function.
    if analyze:
        data, save_master = analize_data(data)
        header_text = np.array(['time', 'pos x', 'pos y', 'angle', 'difftime',
                            'diffposx', 'diffposy', 'diffangl', 'difflong',
                            'relspeed'])
    else:
        header_text = np.array(['time', 'pos x', 'pos y', 'angle'])
        save_master = False
    full_data = np.vstack([header_text, data])
    # Name of the output file for the poses historic values.
    datestamp = "{}".format(time.strftime("%d_%m_%Y_%H%M"))
    filename = "datatemp/{}-L{}-R{}".format(datestamp, sp_left, sp_right)
    name_spreadsheet = "{}.xlsx".format(filename)
    name_txt = "{}.txt".format(filename)
    #Call to save data in spreadsheet.
    data2spreadsheet(header_text, full_data, name_spreadsheet)
    #Header for numpy savetxt.
    header_numpy = ''
    cols = header_text.shape[0]
    for x in range (0, cols):
        element = header_text[x]
        element = '%9s' % (element,)
        header_numpy = '{}{}\t'.format(header_numpy, element)
    #Call to save data in textfile.
    np.savetxt(name_txt, data, delimiter='\t', fmt='%9.2f',
               header=header_numpy, comments='')
    #Save data to masterfile.
    if save_master:
        #The average speed data is in the last row and last column.
        last_row = data.shape[0]-1
        last_col = data.shape[1]-1
        avg_speed = round(float(data[last_row,last_col]),2)
        save2master_xlsx(avg_speed, sp_left, sp_right)
        save2master_txt(avg_speed, sp_left, sp_right)
#    data2textfile(header_text, data, filename_textfile)

def analize_data(data):
    """
    Receives poses and time of matrix to analyze.

    Different time and position values are calculated between two data points.
    From these, the displaced length and the average speed of UGV are calculated.

    :param data: Matrix of floats32 with data.
    """
    rows = data.shape[0]
    cols = data.shape[1]
    #Differential data matrix: current data minus previous data.
    diff_data = np.zeros_like(data)
    #Vector differential length displaced.
    diff_length =  np.zeros(rows)
    #Vector differential speed.
    diff_speed = np.zeros(rows)
    for x in range(2, rows):
        for y in range(0, cols):
            diff_data[x,y] = data[x,y]-data[x-1,y]
        diff_length[x] = pow(diff_data[x,1],2) + pow(diff_data[x,2],2)
        #Increase or decrease of displacement.
        if diff_data[x,1]>0:
            diff_length[x] = math.sqrt(diff_length[x])
        else:
            diff_length[x] = -math.sqrt(diff_length[x])
        #Prevents division between zero.
        if diff_data[x,0] != 0:
            diff_speed[x] = diff_length[x] / diff_data[x,0]
    #Complete data matrix with new data.
    diff_data = np.insert(diff_data, 4, diff_length, axis=1)
    diff_data = np.insert(diff_data, 5, diff_speed, axis=1)
    data = np.hstack([data, diff_data])
    #Vector sum of columns of data.
    sum_data = diff_data.sum(axis=0)
    a = np.array([0, 0, 0, 0])
    sum_data = np.hstack([a, sum_data])
    data = np.vstack([data, sum_data])
    #If you want to save to master file boolean True.
    save_master = True

    return data, save_master

def data2spreadsheet(header, data, filename_spreadsheet):
    """
    Receives poses and time, and saves them in a spreadsheet.

    :param data: contains data to save in spreadsheet.
    :param filename_spreadsheet: name of spreadsheet where the data
     will be saved.
    """
    try:
        wb = load_workbook(filename_spreadsheet)
    except:
        wb = Workbook()
    ws = wb.active
    rows = data.shape[0]
    cols = data.shape[1]
    #Write in spreadsheet the headboard
    for y in range (0, cols):
        ws.cell(column=y+1, row=1, value=header[y])
    #Write in spreadsheet the data
    for x in range(1, rows):
        for y in range(0, cols):
            element = round(float(data[x,y]),2)
            ws.cell(column=y+1, row=x+1, value=element)
    wb.save(filename_spreadsheet)

#TODO NOT USED
def data2textfile(headboard, data, filename_textfile):
    """
    Receives poses and time, and saves them in a textfile.

    :param data: contains data to save in a textfile.
    :param filename_textfile: name of textfile where the data will
     be saved.
    """
    text = ''
    with open(filename_textfile, 'a') as outfile:
        rows = data.shape[0]
        cols = data.shape[1]
        #Write in spreadsheet the headboard
        for y in range (0, cols):
            text = text + "{}".format(headboard[y])
            add_space = True

        text = text + "\n"
        for x in range(1, rows):
            for y in range(0, cols):
                element = float("{0:.2f}".format(data[x,y]))
                text = text + "{} \t\t\t".format(element)
            text = text + "\n"
        outfile.write(text)

def save2master_xlsx(avg_speed, sp_left, sp_right):
    """
    Average speed, setpoint left and right wheels are saved in the same spreadsheet.

    :param avg_speed: float32 with average speed.
    :param sp_left: integer between 0 and 255 of setpoint left speed.
    :param sp_right: integer between 0 and 255 of setpoint right speed.
    """
    try:
        wb = load_workbook("datatemp/masterfile.xlsx")
    except:
        wb = Workbook()
    ws = wb.active
    #Next empty row search.
    row = 1
    written_row = True
    while written_row:
        element = ws.cell(column=1, row=row).value
        if element == None:
            written_row = False
        else:
            row +=1
    #Write data in empty row.
    ws.cell(column=1, row=row, value=avg_speed)
    ws.cell(column=2, row=row, value=sp_left)
    ws.cell(column=3, row=row, value=sp_right)
    wb.save("datatemp/masterfile.xlsx")

def save2master_txt(avg_speed, sp_left, sp_right):
    """
    Average speed, setpoint left and right wheels are saved in the same textfile.

    :param avg_speed: float32 with average speed.
    :param sp_left: integer between 0 and 255 of setpoint left speed.
    :param sp_right: integer between 0 and 255 of setpoint right speed.
    """
    text = ''
    with open("datatemp/masterfile.txt", 'a') as outfile:
    #TODO improve format
        text = text + "{}\t\t\t{}\t\t{}\t\t\n".format(avg_speed, sp_left, 
                                                      sp_right)
        outfile.write(text)